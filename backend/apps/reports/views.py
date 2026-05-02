"""
Reports Views - Excel and PDF report generation
"""
import io
import calendar
from datetime import date

from django.http import HttpResponse
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.common.permissions import IsOwnerOrCoordinator
from apps.trips.models import Trip
from apps.expenses.models import Expense
from apps.payments.models import VehicleSettlement


def _month_label(year, month):
    return f"{calendar.month_name[month]} {year}"


def _make_excel_styles():
    """Return a dict of reusable openpyxl style objects."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'right': Alignment(horizontal='right', vertical='center'),
        'thin': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        ),
    }


class MonthlyMISReportView(APIView):
    """Monthly MIS Excel - regular trips sheet + adhoc trips sheet."""
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrCoordinator]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        year_str = request.query_params.get('year')
        month_str = request.query_params.get('month')
        vehicle_id = request.query_params.get('vehicle_id')

        if not year_str or not month_str:
            return Response(
                {'success': False, 'message': 'year and month are required'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            year, month = int(year_str), int(month_str)
            if not 1 <= month <= 12:
                raise ValueError
        except ValueError:
            return Response(
                {'success': False, 'message': 'Invalid year or month'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = Trip.objects.select_related('driver__user', 'vehicle').filter(
            trip_date__year=year,
            trip_date__month=month,
        ).order_by('vehicle__vehicle_number', 'trip_date')

        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        regular = list(qs.filter(trip_category='regular'))
        adhoc = list(qs.filter(trip_category='adhoc'))

        s = _make_excel_styles()
        HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
        HDR_FILL = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        TOT_FONT = Font(bold=True, size=10)
        TOT_FILL = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid')
        EVEN_FILL = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

        COLS = [
            ('Date', 12),
            ('Driver Name', 20),
            ('Vehicle No.', 14),
            ('Store 1', 22),
            ('KM 1 (One-way)', 15),
            ('KM 1 (Round)', 13),
            ('Store 2', 22),
            ('KM 2 (One-way)', 15),
            ('KM 2 (Round)', 13),
            ('Total KM', 10),
            ('Dispatch Time', 13),
            ('Status', 11),
            ('Remarks', 26),
        ]
        label = _month_label(year, month)

        def build_sheet(ws, trips, sheet_title):
            ws.title = sheet_title
            ws.freeze_panes = 'A3'

            last_col = get_column_letter(len(COLS))
            ws.merge_cells(f'A1:{last_col}1')
            ws['A1'].value = f"JJR Logistics — {sheet_title} | {label}"
            ws['A1'].font = Font(bold=True, size=13, color='1E3A8A')
            ws['A1'].alignment = s['center']
            ws.row_dimensions[1].height = 26

            for ci, (name, width) in enumerate(COLS, 1):
                c = ws.cell(row=2, column=ci, value=name)
                c.font = HDR_FONT
                c.fill = HDR_FILL
                c.alignment = s['center']
                c.border = s['thin']
                ws.column_dimensions[get_column_letter(ci)].width = width
            ws.row_dimensions[2].height = 22

            total_km_sum = 0
            last_row = 2

            for ri, trip in enumerate(trips, 3):
                last_row = ri
                driver_name = ''
                if trip.driver and trip.driver.user:
                    u = trip.driver.user
                    driver_name = u.get_full_name() or str(u.phone or '')
                vehicle_num = trip.vehicle.vehicle_number if trip.vehicle else ''
                dispatch_str = trip.dispatch_time_1.strftime('%H:%M') if trip.dispatch_time_1 else ''
                km1_1w = float(trip.one_way_km_1 or 0)
                km2_1w = float(trip.one_way_km_2 or 0)
                total_km = float(trip.total_km or 0)
                total_km_sum += total_km

                row_data = [
                    trip.trip_date.strftime('%d-%m-%Y'),
                    driver_name,
                    vehicle_num,
                    trip.store_name_1 or '—',
                    km1_1w or '—',
                    km1_1w * 2 or '—',
                    trip.store_name_2 or '—',
                    km2_1w or '—',
                    km2_1w * 2 or '—',
                    total_km,
                    dispatch_str or '—',
                    trip.get_status_display(),
                    trip.remarks or '',
                ]
                fill = EVEN_FILL if ri % 2 == 0 else None

                for ci, val in enumerate(row_data, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = s['thin']
                    if fill:
                        c.fill = fill
                    if ci == 1:
                        c.alignment = s['center']
                    elif ci in (5, 6, 8, 9, 10) and isinstance(val, float):
                        c.alignment = s['right']
                        c.number_format = '0.0'

            # Totals row
            tr = last_row + 1
            for ci in range(1, len(COLS) + 1):
                c = ws.cell(row=tr, column=ci)
                c.border = s['thin']
                c.fill = TOT_FILL
                c.font = TOT_FONT
            ws.cell(row=tr, column=1).value = 'TOTAL'
            ws.cell(row=tr, column=1).alignment = s['center']
            tc = ws.cell(row=tr, column=10)
            tc.value = round(total_km_sum, 1)
            tc.number_format = '0.0'
            tc.alignment = s['right']

            summary_row = tr + 2
            ws.cell(row=summary_row, column=1).value = f"Total Trips: {len(trips)}"
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=10, color='374151')

        wb = openpyxl.Workbook()
        build_sheet(wb.active, regular, 'Regular Trips')
        build_sheet(wb.create_sheet(), adhoc, 'Adhoc Trips')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"MIS_{month:02d}_{year}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp


class ExpenseReportView(APIView):
    """Expense report Excel with optional filters."""
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrCoordinator]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from django.db.models import Sum

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        vehicle_id = request.query_params.get('vehicle_id')
        expense_type = request.query_params.get('expense_type')

        qs = Expense.objects.select_related('vehicle', 'created_by').order_by(
            'expense_date', 'expense_type'
        )
        if start_date:
            qs = qs.filter(expense_date__gte=start_date)
        if end_date:
            qs = qs.filter(expense_date__lte=end_date)
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)
        if expense_type:
            qs = qs.filter(expense_type=expense_type)

        expenses = list(qs)

        s = _make_excel_styles()
        HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
        HDR_FILL = PatternFill(start_color='065F46', end_color='065F46', fill_type='solid')
        TOT_FONT = Font(bold=True, size=10)
        TOT_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        EVEN_FILL = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')

        COLS = [
            ('Date', 12),
            ('Vehicle', 16),
            ('Expense Type', 18),
            ('Amount (₹)', 14),
            ('Payment Mode', 15),
            ('Paid To', 20),
            ('Remarks', 30),
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Expenses'
        ws.freeze_panes = 'A3'

        date_range = f"{start_date or 'All'} to {end_date or 'All'}"
        last_col = get_column_letter(len(COLS))
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].value = f"JJR Logistics — Expense Report | {date_range}"
        ws['A1'].font = Font(bold=True, size=13, color='065F46')
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 26

        for ci, (name, width) in enumerate(COLS, 1):
            c = ws.cell(row=2, column=ci, value=name)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = s['center']
            c.border = s['thin']
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[2].height = 22

        total_amount = 0

        for ri, exp in enumerate(expenses, 3):
            vehicle_num = exp.vehicle.vehicle_number if exp.vehicle else '—'
            paid_to = exp.paid_to_name or exp.paid_to_number or '—'
            amount = float(exp.amount)
            total_amount += amount

            row_data = [
                exp.expense_date.strftime('%d-%m-%Y'),
                vehicle_num,
                exp.get_expense_type_display(),
                amount,
                exp.get_payment_mode_display(),
                paid_to,
                exp.remarks or '',
            ]
            fill = EVEN_FILL if ri % 2 == 0 else None

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = s['thin']
                if fill:
                    c.fill = fill
                if ci == 1:
                    c.alignment = s['center']
                elif ci == 4:
                    c.alignment = s['right']
                    c.number_format = '#,##0.00'

        # Totals row
        tr = len(expenses) + 3
        for ci in range(1, len(COLS) + 1):
            c = ws.cell(row=tr, column=ci)
            c.border = s['thin']
            c.fill = TOT_FILL
            c.font = TOT_FONT
        ws.cell(row=tr, column=1).value = 'TOTAL'
        ws.cell(row=tr, column=1).alignment = s['center']
        tc = ws.cell(row=tr, column=4)
        tc.value = round(total_amount, 2)
        tc.number_format = '#,##0.00'
        tc.alignment = s['right']

        # Breakdown by type
        summary_start = tr + 2
        ws.cell(row=summary_start, column=1).value = 'Breakdown by Type'
        ws.cell(row=summary_start, column=1).font = Font(bold=True, size=10)

        type_totals = qs.values('expense_type').annotate(total=Sum('amount')).order_by('expense_type')
        for i, row in enumerate(type_totals, 1):
            ws.cell(row=summary_start + i, column=1).value = (
                row['expense_type'].replace('_', ' ').title()
            )
            c = ws.cell(row=summary_start + i, column=2)
            c.value = float(row['total'])
            c.number_format = '#,##0.00'
            c.alignment = s['right']

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        suffix = f"{start_date or 'all'}_{end_date or 'all'}"
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="Expenses_{suffix}.xlsx"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp


class PaymentSummaryReportView(APIView):
    """Vehicle settlement summary as Excel (default) or PDF (format=pdf)."""
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrCoordinator]

    def get(self, request):
        year_str = request.query_params.get('year')
        month_str = request.query_params.get('month')
        fmt = request.query_params.get('format', 'excel')

        if not year_str or not month_str:
            return Response(
                {'success': False, 'message': 'year and month are required'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            year, month = int(year_str), int(month_str)
            if not 1 <= month <= 12:
                raise ValueError
        except ValueError:
            return Response(
                {'success': False, 'message': 'Invalid year or month'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        month_start = date(year, month, 1)
        settlements = list(
            VehicleSettlement.objects.select_related('vehicle', 'paid_by')
            .filter(month_year=month_start)
            .order_by('vehicle__vehicle_number')
        )
        label = _month_label(year, month)

        if fmt == 'pdf':
            return self._build_pdf(settlements, label, year, month)
        return self._build_excel(settlements, label, year, month)

    def _build_excel(self, settlements, label, year, month):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        s = _make_excel_styles()
        HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
        HDR_FILL = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        TOT_FONT = Font(bold=True, size=10)
        TOT_FILL = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
        EVEN_FILL = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')

        COLS = [
            ('Vehicle', 16),
            ('Working Days', 13),
            ('Total KM', 10),
            ('Base Amt (₹)', 14),
            ('Penalty (₹)', 13),
            ('Extra KM (₹)', 13),
            ('Gross Amt (₹)', 14),
            ('Total Expenses (₹)', 18),
            ('Carry Fwd (₹)', 14),
            ('Balance Payable (₹)', 18),
            ('Status', 12),
            ('Paid Amt (₹)', 13),
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Settlement Summary'
        ws.freeze_panes = 'A3'

        last_col = get_column_letter(len(COLS))
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].value = f"JJR Logistics — Settlement Summary | {label}"
        ws['A1'].font = Font(bold=True, size=13, color='7C3AED')
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 26

        for ci, (name, width) in enumerate(COLS, 1):
            c = ws.cell(row=2, column=ci, value=name)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = s['center']
            c.border = s['thin']
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[2].height = 28

        total_balance = 0

        for ri, stl in enumerate(settlements, 3):
            balance = float(stl.balance_payable)
            total_balance += balance

            row_data = [
                stl.vehicle.vehicle_number if stl.vehicle else '—',
                stl.working_days,
                float(stl.total_km),
                float(stl.base_amount),
                float(stl.absent_penalty_amount),
                float(stl.extra_km_amount),
                float(stl.gross_amount),
                float(stl.total_expenses),
                float(stl.carry_forward_from_previous),
                balance,
                stl.get_status_display(),
                float(stl.paid_amount),
            ]
            fill = EVEN_FILL if ri % 2 == 0 else None

            money_cols = {4, 5, 6, 7, 8, 9, 10, 12}
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = s['thin']
                if fill:
                    c.fill = fill
                if ci in money_cols:
                    c.alignment = s['right']
                    c.number_format = '#,##0.00'
                elif ci in (2, 3):
                    c.alignment = s['right']
                elif ci == 11:
                    c.alignment = s['center']

        # Totals
        tr = len(settlements) + 3
        for ci in range(1, len(COLS) + 1):
            c = ws.cell(row=tr, column=ci)
            c.border = s['thin']
            c.fill = TOT_FILL
            c.font = TOT_FONT
        ws.cell(row=tr, column=1).value = 'TOTAL PAYABLE'
        ws.cell(row=tr, column=1).alignment = s['center']
        tc = ws.cell(row=tr, column=10)
        tc.value = round(total_balance, 2)
        tc.number_format = '#,##0.00'
        tc.alignment = s['right']

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"Settlements_{month:02d}_{year}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp

    def _build_pdf(self, settlements, label, year, month):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        purple = colors.HexColor('#7C3AED')
        light_purple = colors.HexColor('#F5F3FF')
        yellow = colors.HexColor('#FEF08A')

        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Title'],
            textColor=purple, fontSize=15, spaceAfter=2,
        )
        sub_style = ParagraphStyle(
            'ReportSub', parent=styles['Normal'],
            textColor=colors.HexColor('#374151'), fontSize=10, spaceAfter=6,
        )

        elements = [
            Paragraph("JJR Logistics — Settlement Summary", title_style),
            Paragraph(label, sub_style),
            Spacer(1, 4 * mm),
        ]

        headers = [
            'Vehicle', 'Working Days', 'Total KM',
            'Base Amt (₹)', 'Gross Amt (₹)', 'Total Exp (₹)',
            'Balance (₹)', 'Status',
        ]
        data = [headers]
        total_balance = 0

        for stl in settlements:
            balance = float(stl.balance_payable)
            total_balance += balance
            data.append([
                stl.vehicle.vehicle_number if stl.vehicle else '—',
                str(stl.working_days),
                f"{float(stl.total_km):.1f}",
                f"{float(stl.base_amount):,.0f}",
                f"{float(stl.gross_amount):,.0f}",
                f"{float(stl.total_expenses):,.0f}",
                f"{balance:,.0f}",
                stl.get_status_display(),
            ])

        data.append(['TOTAL', '', '', '', '', '', f"{total_balance:,.0f}", ''])

        col_widths = [28*mm, 22*mm, 18*mm, 24*mm, 24*mm, 24*mm, 24*mm, 18*mm]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), purple),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (1, 1), (6, -2), 'RIGHT'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, light_purple]),
            ('BACKGROUND', (0, -1), (-1, -1), yellow),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('ALIGN', (6, -1), (6, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHT', (0, 0), (-1, 0), 16),
            ('ROWHEIGHT', (0, 1), (-1, -1), 14),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 5 * mm))
        elements.append(Paragraph(
            f"Generated: {date.today().strftime('%d %B %Y')} | "
            f"Records: {len(settlements)} | "
            f"Total Balance Payable: ₹{total_balance:,.0f}",
            styles['Normal'],
        ))

        doc.build(elements)
        buf.seek(0)

        fname = f"Settlements_{month:02d}_{year}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp
